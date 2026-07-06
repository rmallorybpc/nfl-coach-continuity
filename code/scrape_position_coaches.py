#!/usr/bin/env python3
"""
Fill columns I-Q (position coaches) of NFL_Coaching_Staff_2013_2026.xlsx
from pro-football-history.com.  FIXED VERSION.

What changed from v1:
  - Uses a normal browser User-Agent (v1 identified as 'python-requests',
    which the site blocked -> zero pages fetched).
  - Sturdier staff parser (matches on visible label text).
  - A SELF-TEST runs first and tells you in seconds whether fetching and
    parsing work, before committing to the full crawl.
  - Writer skips non-data rows (the 'AFC Team' crash).
  - Uses its own checkpoint file, so any old .scrape_checkpoint.json is ignored.

SETUP (VS Code terminal, Windows):
    py -m pip install requests beautifulsoup4 openpyxl
RUN:
    py scrape_position_coaches_v2.py
"""

import json, re, time, random, argparse, sys
from pathlib import Path
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

SHEET = "NFL Coaching Staff 2013-2026"
SEED = "https://pro-football-history.com/season/910/2013-seattle-seahawks-schedule"
BASE = "https://pro-football-history.com"
CKPT = Path(".scrape_checkpoint_v2.json")
AUDIT = Path("position_coach_audit.csv")

LABEL_TO_COL = {
    "Quarterbacks Coach": "QB Coach",
    "Running Backs Coach": "RB Coach",
    "Wide Receivers Coach": "WR Coach",
    "Tight Ends Coach": "TE Coach",
    "Offensive Line Coach": "OL Coach",
    "Defensive Line Coach": "DL Coach",
    "Linebackers Coach": "LB Coach",
    "Secondary Coach": "DB/Secondary Coach",
    "Defensive Backs Coach": "DB/Secondary Coach",
    "Special Teams Coordinator": "Special Teams Coach",
    "Special Teams Coach": "Special Teams Coach",
}
TARGET_COLS = ["QB Coach", "RB Coach", "WR Coach", "TE Coach", "OL Coach",
               "DL Coach", "LB Coach", "DB/Secondary Coach", "Special Teams Coach"]

HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}
SLUG_RE = re.compile(r"/season/\d+/(\d{4})-([a-z0-9-]+?)-schedule")
FRANCHISE_RE = re.compile(r"/franchise/\d+/[a-z0-9-]+-coaches")


def norm_team(s):
    s = str(s).lower().replace(".", "").replace("-", " ")
    return re.sub(r"\s+", " ", s).strip()


def slug_to_team(slug):
    return slug.replace("-", " ").title().replace("St ", "St. ")


def get(url, tries=4):
    last = None
    for i in range(tries):
        try:
            r = requests.get(url, headers=HEADERS, timeout=25)
            last = r.status_code
            if r.status_code == 200:
                return r.text, 200
            if r.status_code in (403, 404, 410):
                return None, r.status_code
        except requests.RequestException as e:
            last = repr(e)
        time.sleep(2 ** i + random.random())
    return None, last


def parse_staff(html):
    soup = BeautifulSoup(html, "html.parser")
    strongs = soup.find_all("strong")
    start = next((i for i, s in enumerate(strongs)
                  if s.get_text(strip=True).lower() == "head coach"), None)
    if start is None:
        return {}, []
    row, multi = {}, []
    for s in strongs[start:]:
        label = s.get_text(strip=True)
        names = []
        for sib in s.next_siblings:
            nm = getattr(sib, "name", None)
            if nm == "strong":
                break
            if nm == "a":
                t = sib.get_text(strip=True)
                if t:
                    names.append(t)
        col = LABEL_TO_COL.get(label)
        if col and names and col not in row:
            row[col] = names[0]
            if len(names) > 1:
                multi.append((col, names))
    return row, multi


def self_test():
    print("SELF-TEST: fetching one page to check access and parsing...")
    html, status = get(SEED)
    if not html:
        print(f"  FAILED to fetch (status/error: {status}).")
        print("  The site is not reachable from this machine, or is blocking the request.")
        print("  Paste this whole output to me and we will adjust.")
        return False
    staff, _ = parse_staff(html)
    got = sum(1 for c in TARGET_COLS if staff.get(c))
    print(f"  fetch OK (200), parsed {got}/9 position-coach cells for 2013 Seattle Seahawks")
    if got >= 7:
        print("  looks good. starting full crawl.\n")
        return True
    print("  fetched the page but could not parse the staff. Paste this output to me.")
    print("  First 400 chars of page:\n", html[:400])
    return False


def crawl(year_lo, year_hi, delay, max_pages):
    results, visited = {}, set()
    if CKPT.exists():
        saved = json.loads(CKPT.read_text())
        results, visited = saved.get("results", {}), set(saved.get("visited", []))
        print(f"resuming: {len(results)} team-seasons already scraped")
    queue, pages, audit = [SEED], 0, []
    while queue and pages < max_pages:
        url = queue.pop(0)
        if url in visited:
            continue
        visited.add(url)
        html, status = get(url)
        pages += 1
        if not html:
            continue
        for href in re.findall(r'href="([^"]+)"', html):
            full = href if href.startswith("http") else BASE + href
            m = SLUG_RE.search(full)
            if m:
                if year_lo <= int(m.group(1)) <= year_hi and full not in visited:
                    queue.append(full)
            elif FRANCHISE_RE.search(full) and full not in visited:
                queue.append(full)
        m = SLUG_RE.search(url)
        if m and year_lo <= int(m.group(1)) <= year_hi:
            yr, team = int(m.group(1)), slug_to_team(m.group(2))
            key = f"{yr}|{norm_team(team)}"
            if key not in results:
                staff, multi = parse_staff(html)
                if staff:
                    results[key] = staff
                    for col, names in multi:
                        audit.append((yr, team, col, " / ".join(names)))
                    print(f"  [{len(results):3}] {yr} {team:24} "
                          f"{sum(1 for c in TARGET_COLS if staff.get(c))}/9")
        if pages % 10 == 0:
            CKPT.write_text(json.dumps({"results": results, "visited": list(visited)}))
        time.sleep(delay + random.random() * 0.5)
    CKPT.write_text(json.dumps({"results": results, "visited": list(visited)}))
    return results, audit


def find_header_row(ws):
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [str(ws.cell(r, c).value).strip() if ws.cell(r, c).value is not None else ""
                for c in range(1, ws.max_column + 1)]
        if "Season" in vals and "Team" in vals:
            return r
    return 3


def write_xlsx(infile, outfile, results):
    wb = load_workbook(infile)
    ws = wb[SHEET]
    hr = find_header_row(ws)
    headers = {ws.cell(hr, c).value: c for c in range(1, ws.max_column + 1)}
    s_col, t_col = headers["Season"], headers["Team"]
    col_idx = {n: headers[n] for n in TARGET_COLS if n in headers}
    filled = missing = skipped = 0
    for r in range(hr + 1, ws.max_row + 1):
        season, team = ws.cell(r, s_col).value, ws.cell(r, t_col).value
        if season is None or team is None:
            continue
        try:
            yr = int(season)
        except (ValueError, TypeError):
            skipped += 1
            continue
        staff = results.get(f"{yr}|{norm_team(team)}")
        if not staff:
            missing += 1
            continue
        for n, idx in col_idx.items():
            if staff.get(n):
                ws.cell(r, idx).value = staff[n]
        filled += 1
    wb.save(outfile)
    print(f"\nwrote {outfile}: {filled} filled, {missing} empty, {skipped} non-data rows skipped")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--infile", default="NFL_Coaching_Staff_2013_2026.xlsx")
    ap.add_argument("--outfile", default="NFL_Coaching_Staff_2013_2026_FILLED.xlsx")
    ap.add_argument("--year-lo", type=int, default=2013)
    ap.add_argument("--year-hi", type=int, default=2026)
    ap.add_argument("--delay", type=float, default=1.5)
    ap.add_argument("--max-pages", type=int, default=2000)
    ap.add_argument("--skip-test", action="store_true")
    a = ap.parse_args()
    if not Path(a.infile).exists():
        sys.exit(f"input not found in this folder: {a.infile}")
    if not a.skip_test and not self_test():
        sys.exit(1)
    results, audit = crawl(a.year_lo, a.year_hi, a.delay, a.max_pages)
    if audit:
        with AUDIT.open("w") as f:
            f.write("season,team,column,names\n")
            for yr, team, col, names in audit:
                f.write(f'{yr},"{team}",{col},"{names}"\n')
        print(f"wrote {AUDIT} ({len(audit)} multi-name cells)")
    write_xlsx(a.infile, a.outfile, results)
    print("done. send me the filled file.")


if __name__ == "__main__":
    main()
